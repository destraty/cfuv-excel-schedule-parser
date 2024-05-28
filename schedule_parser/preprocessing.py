import asyncio
import re
import os

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from common import ROOT_DIR
from schedule_parser.group_fixer import bad_group_name_fixer
from schedule_parser.xls2xlsx import xls2xlsx
from schedule_parser.parser import parse_data
from schedule_parser.post_processing import working_with_cells
from schedule_parser.execution_function import calculate_execution_time


pattern = re.compile(
    r'([А-Яа-я][А-Яа-я]?[А-Яа-я]?-[а-я]+.*?-[а-я]+.*?[0-9][0-9][0-9].*|С.[А-Я]+.-[а-я]+-[а-я]+-[0-9]+|[А-Яа-я][А-Яа-я]?[0-9][0-9][0-9].*|ЦК-[0-9][0-9][0-9]?|[А-Я]{1,4}.?[А-Я]?.?-[а-я]-[а-я]-[0-9]{3})')


def adjust_rows_in_excel_files(sheet_obj: Worksheet) -> int:
    """
    Возвращает int - значение сдвига расписания по вертикали
    -value = кол-во лишних строк
    +value = кол-во недостающих строк
    0 = все подходит

    :return: +-value | 0 - vertical schedule offset value
    """
    # workbook = load_workbook(filename=file_path, read_only=True)
    target_row_index = None
    for i, row in enumerate(sheet_obj.iter_rows(values_only=True), start=1):
        if any("Дни недели" in str(cell) for cell in row if cell is not None):
            target_row_index = i
            break
    if not target_row_index:
        return 0
    # Если нашли строку с "Дни недели" и она не на 3 позиции
    if target_row_index is not None and target_row_index != 3:
        # Вычисляем, сколько строк нужно добавить или удалить
        diff = 3 - target_row_index
        return diff
    elif target_row_index == 3:
        return 0


@calculate_execution_time
def preprocessing(cell_index: int = 4) -> list[list[str | list[str | list[str]]]]:
    asyncio.run(parse_data())
    # Преобразовываем xls в xlsx
    xls2xlsx(f'{ROOT_DIR}/schedule/')
    bad_group_name_fixer()
    valid_results, invalid_results = [], []
    files_names = [f for f in os.listdir(f'{ROOT_DIR}/schedule')]

    # читаем каждый файл и добавляем названия листов и группы из строки 4 каждого листа
    for filename in files_names:
        workbook = openpyxl.load_workbook(f'{ROOT_DIR}/schedule/{filename}', read_only=True)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            groups = set()
            c_a = cell_index - adjust_rows_in_excel_files(worksheet)
            try:
                worksheet[c_a]
            except IndexError:
                break
            for ind in [c_a, c_a - 1]:
                for cell in worksheet[ind]:
                    if cell.value is not None and pattern.findall(str(cell.value).replace(
                            ' ', '')):
                        group_name = str(cell.value).replace(
                            'Подгруппа', '').replace(
                            'группа', '').replace(
                            ' ', '').strip()

                        if not any(group_name in group for group in groups):
                            groups.add(cell.value)
            valid_results.append([filename, [sheet_name, list(groups)]])

    invalid_results.extend([res for res in valid_results if res[1][1] == []])
    valid_results = [res for res in valid_results if res[1][1] != []]

    for res in valid_results:
        print(res)

    print('=' * 60)
    for res in invalid_results:
        print(res)
    print(f'Валидных: {len(valid_results)} из {len(valid_results) + len(invalid_results)}')

    working_with_cells()
    return valid_results

