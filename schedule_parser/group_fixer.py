import os
import re
import shutil

from openpyxl.reader.excel import load_workbook

from schedule_parser.execution_function import calculate_execution_time
from common import ROOT_DIR


@calculate_execution_time
def bad_group_name_fixer() -> None:
    """
    Пластырь для файлов, которые имеют подгруппы в не правильном формате.
    Пример: вместо ПИ-б-о-211(1) просто стоит 1 или (1) или 1 подгруппа
    :return: None
    """
    # Путь к папке с файлами
    folder_path = f"{ROOT_DIR}/schedule/need_handle"

    # Шаблон для поиска группы в файле
    pattern = re.compile(
        r'([А-Яа-я][А-Яа-я]?[А-Яа-я]?-[а-я]+.*?-[а-я]+.*?[0-9][0-9][0-9].*|С.[А-Я]+.-[а-я]+-[а-я]+-[0-9]+|[А-Яа-я][А-Яа-я]? [0-9][0-9][0-9].*|ЦК-[0-9][0-9][0-9]?)')

    # Получение списка файлов в папке
    files = [f for f in os.listdir(folder_path)]

    # Обработка каждого файла
    for file in files:
        file_path = os.path.join(folder_path, file)

        # Загрузка файла xlsx
        workbook = load_workbook(file_path)

        # Обработка каждого листа в файле
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Проверка наличия 4 строки
            if sheet.max_row >= 4:
                # Получение 4 строки
                row = sheet[3]
                # Разъединение ячеек в 4 строке
                for col in range(sheet.max_column):
                    try:
                        sheet.unmerge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

                    except:
                        pass
                    try:
                        sheet.unmerge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
                    except:
                        pass
                    try:
                        sheet.unmerge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
                    except:
                        pass
                    try:
                        sheet.unmerge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 2)
                    except:
                        pass
                    # try:
                    #     sheet.unmerge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
                    # except:
                    #     pass
                if "ОФО" in str(file):
                    for i, cell in enumerate(row):
                        if cell.value is not None and i > 1 and (pattern.match(str(cell.value))):
                            # Замена значений ячеек
                            sheet.cell(row=4, column=i + 1).value = str(
                                cell.value.strip()) + "(1)"
                            sheet.cell(row=4, column=i + 2).value = str(
                                cell.value).strip() + "(2)"
                else:
                    for i, cell in enumerate(sheet[4]):
                        if cell.value is not None and i > 1 and (pattern.match(
                                str(cell.value))) and "(1)" not in cell.value and "(2)" not in cell.value and "(3)" not in cell.value:
                            saved_val = cell.value
                            print(f'detected: {cell.value}')
                            # Замена значений ячеек
                            if sheet.cell(row=5, column=i + 1).value is not None:
                                print(
                                    f'replaced {sheet.cell(row=5, column=i + 1).value}to {str(cell.value.strip()) + "(n)"}')
                                sheet.cell(row=4, column=i + 1).value = str(
                                    saved_val.strip()) + "(1)"
                                sheet.cell(row=4, column=i + 2).value = str(
                                    saved_val).strip() + "(2)"
                                if "3" in str(sheet.cell(row=5, column=i + 3).value):
                                    sheet.cell(row=4, column=i + 3).value = str(
                                        saved_val).strip() + "(3)"

        # Сохранение изменений в файле xlsx
        workbook.save(file_path)
    # Путь к исходной папке
    source_folder = f'{ROOT_DIR}/schedule/need_handle'

    # Путь к целевой папке
    target_folder = f'{ROOT_DIR}/schedule'

    # Получаем список всех файлов в исходной папке
    files = os.listdir(source_folder)
    # Перемещаем каждый файл из исходной папки в целевую папку
    for file_name in files:
        # Полный путь к исходному файлу
        source_file = os.path.join(source_folder, file_name)
        # Полный путь к целевому файлу
        target_file = os.path.join(target_folder, file_name)
        # Перемещаем файл
        shutil.move(source_file, target_file)

    os.rmdir(f'{ROOT_DIR}/schedule/need_handle')

if __name__ == "__main__":
    bad_group_name_fixer()
