import os
import concurrent.futures
import multiprocessing

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from schedule_parser.execution_function import calculate_execution_time
from concurrent.futures import ThreadPoolExecutor
from common import ROOT_DIR

folder_path = f"{ROOT_DIR}/schedule"


def process_sheet(workbook: Workbook, sheet_name: str) -> None:
    """
    Функция обработки листа. Разъединяет только объединенные по строкам ячейки.
    :param workbook: Объект открытой xlsx книги
    :param sheet_name: Имя листа книги для обработки
    :return: Лист книги с разъединенными ячейками
    """
    # Получаем текущий лист
    worksheet = workbook[sheet_name]

    # Создаем копию множества объединенных ячеек
    merged_cells = set(worksheet.merged_cells.ranges)

    # Проходимся по каждой объединенной ячейке
    for merged_cell in merged_cells:
        # Проверяем, что ячейка объединена только по строкам
        if merged_cell.min_col == merged_cell.max_col:
            # Разъединяем ячейку
            worksheet.unmerge_cells(str(merged_cell))


@calculate_execution_time
def padding_a_file_with_spaces(file_name: str) -> None:
    """
    Решающая в дальнейшем парсинге функция.
    Для корректной отработки функции pandas.DataFrame.ffill заполняем все пустые не объединенные ячейки пробелами
    :param file_name: Имя документа в виде пути
    :return: Обработанный файл
    """
    # Формируем полный путь к файлу
    file_path = os.path.join(folder_path, file_name)

    # Открываем файл
    workbook = load_workbook(file_path)
    wb_sheet_names = workbook.sheetnames
    with ThreadPoolExecutor(max_workers=len(wb_sheet_names)) as executor:
        # Создаем список задач для каждого листа
        tasks = [executor.submit(process_sheet, workbook, sheet) for sheet in wb_sheet_names]

        # Дожидаемся завершения всех задач
        for task in tasks:
            task.result()

    # Проходим по всем листам в файле
    for sheet in wb_sheet_names:
        # Выбираем текущий лист
        current_sheet = workbook[sheet]
        # Создаем множество для хранения координат объединенных ячеек
        merged_cells = set()

        # Получаем список объединенных диапазонов
        for merged_range in current_sheet.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_cells.add((row, col))

        # Проходим по всем ячейкам
        for row in current_sheet.iter_rows():
            for cell in row:
                # Проверяем, является ли ячейка пустой и не объединенной
                if cell.value is None and (cell.row, cell.column) not in merged_cells:
                    # Записываем пробел в нее
                    cell.value = ' '

    # Сохраняем изменения
    workbook.save(file_path)


@calculate_execution_time
def working_with_cells():
    # Получаем список всех файлов в папке
    files = os.listdir(folder_path)

    # Создаем пул потоков с числом потоков равным числу потоков процессора
    with concurrent.futures.ThreadPoolExecutor(max_workers=multiprocessing.cpu_count()) as executor:
        # Запускаем обработку каждого файла в отдельном потоке
        executor.map(padding_a_file_with_spaces, files)


# if __name__ == "__main__":
#     working_with_cells()
